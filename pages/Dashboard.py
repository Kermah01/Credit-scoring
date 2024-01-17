import streamlit as st
import pandas as pd
import plotly.express as px
import openpyxl
from openpyxl import load_workbook
import numpy as np
from streamlit_extras.metric_cards import style_metric_cards # beautify metric card with css
import plotly.graph_objects as go
import nltk
from nltk import word_tokenize, ngrams
from collections import Counter
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import re
from unidecode import unidecode
import streamlit_gsheets
from streamlit_gsheets import GSheetsConnection
from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)

st.set_page_config(layout="wide")
st.sidebar.image('logo.jpg', use_column_width='always')
# Chargement des données
data = openpyxl.load_workbook("Copie de BD finale.xlsx")
datas = data.active
donnees = []
for ligne in datas.iter_rows(values_only=True):
    donnees.append(list(ligne))
en_tetes = donnees[0]
donnees = donnees[1:]
df = pd.DataFrame(donnees, columns=en_tetes)
df.set_index(en_tetes[0], inplace=True)
order_of_months = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
order_of_years = ['2006', '2007', '2008', '2009', '2010', '2011', '2012', '2013', '2014', '2015', '2016', '2017','2018','2019','2020','2021','2022']
def transf_df(df):
    df['AGE'] = df['AGE'].astype(int)
    df['MONTANT SOLLICITE'] = df['MONTANT SOLLICITE'].astype(int)
    df["TAUX D'INTERET"].replace('null', 0, inplace=True)
    df["TAUX D'INTERET"] = df["TAUX D'INTERET"].astype(float)
    df['DATDEP'] = pd.to_datetime(df['DATDEP'], errors='coerce' )
    df["DERNDAT"] = pd.to_datetime(df["DERNDAT"], errors='coerce')
    df.drop(columns=['NOM'], inplace=True)
    d=df.groupby('AGENCES')['STATUT'].sum()
    # Thème des graphiques
    df['AGENCES']=df['AGENCES'].replace('AGENCE- DEUX PLATEAUX LATRILLE','AGENCE- 2 PLATEAUX RUE DES JARDINS' )
    df["MOIS D'OCTROI DU CREDIT"] = df["DATDEP"].dt.month

    # Créez une nouvelle colonne pour l'année
    df["ANNEE D'OCTROI DU CREDIT"] = df["DATDEP"].dt.year
    df["MOIS D'OCTROI DU CREDIT"] = df["MOIS D'OCTROI DU CREDIT"].astype(str)

    df["MOIS D'OCTROI DU CREDIT"]=df["MOIS D'OCTROI DU CREDIT"].map({"1":"Janvier","2":"Février","3":"Mars","4":"Avril","5":"Mai","6":"Juin","7":"Juillet","8":"Août","9":"Septembre","10":"Octobre","11":"Novembre","12":"Décembre"})
    
    df["MOIS D'OCTROI DU CREDIT"] = pd.Categorical(df["MOIS D'OCTROI DU CREDIT"], categories=order_of_months, ordered=True)
    
    return df

page_bg_img = f"""
<style>
[data-testid="stAppViewContainer"] > .main {{
background-image: url(https://static.vecteezy.com/ti/vecteur-libre/p1/20530242-abstrait-arriere-plan-vague-doubler-violet-vague-colore-lignes-neon-lumiere-abstrait-fond-d-ecran-numerique-abstrait-3d-technologie-vague-effet-embrase-lignes-vague-arriere-plan-vectoriel.jpg);
background-size: cover;
background-position: center;
background-repeat: no-repeat;
background-attachment: no-fixed;
height: 100vh;
margin: 0;
display: flex;


}}
[data-testid="stSidebar"] {{
    background-color: #000 !important;  /* Fond noir */
    border: 2px solid #ff0000 !important;  /* Bordure rouge */
    border-radius: 10px;  /* Coins arrondis */
    margin-top: -30px;  /* Ajuster la position vers le haut */
    position: relative;
    z-index: 1;  /* S'assurer que la barre latérale est au-dessus du contenu */
    padding: 10px;
}}

[data-testid="stHeader"] {{
background: rgba(0, 0, 0, 0);
color: white;
}}
[data-testid="stToolbar"] {{
right: 2rem;
}}
</style>
"""

st.markdown(
    """
    <style>
        body {
            color: white;
        }
    </style>
    """,
    unsafe_allow_html=True
)

st.markdown('<div style="text-align:center;width:100%;"><h1 style="color:white;background-color:black;border:red;border-style:solid;border-radius:5px;">TABLEAU DE BORD INTERACTIF</h1></div>', unsafe_allow_html=True)

colors = px.colors.sequential.Rainbow_r
colors.extend(px.colors.sequential.Agsunset)
colors.extend(px.colors.sequential.Aggrnyl)
st.write("\n")
st.write("\n")
# Section des graphiques sommaires
st.markdown(page_bg_img, unsafe_allow_html=True)

st.header("VISUALISATION DE LA BASE DE DONNEES",divider="rainbow" )
def filter_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds a UI on top of a dataframe to let viewers filter columns

    Args:
        df (pd.DataFrame): Original dataframe

    Returns:
        pd.DataFrame: Filtered dataframe
    """
    modify = st.checkbox("AJOUTEZ UN FILTRE")
    

    if not modify:
        return df

    df = df.copy()

    # Try to convert datetimes into a standard format (datetime, no timezone)
    for col in df.columns:
        if is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass

        if is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    with modification_container:
        to_filter_columns = st.multiselect("Choisissez les variables que vous souhaitez utiliser comme filtre", df.columns)
        for column in to_filter_columns:
            left, right = st.columns((1, 20))
            # Treat columns with < 10 unique values as categorical
            int_columns = df.select_dtypes(include="int").columns
            float_columns = df.select_dtypes(include="float").columns

            if is_numeric_dtype(df[column]) :
                _min = int(df[column].min())
                _max = int(df[column].max())
                user_num_input = right.slider(
                    f"Valeurs de {column}",
                    min_value=_min,
                    max_value=_max,
                    value=(_min, _max),
                )
                df = df[df[column].between(*user_num_input)]
            elif is_datetime64_any_dtype(df[column]):
                user_date_input = right.date_input(
                    f"Valeur de {column}",
                    value=(
                        df[column].min(),
                        df[column].max(),
                    ),
                )
                if len(user_date_input) == 2:
                    user_date_input = tuple(map(pd.to_datetime, user_date_input))
                    start_date, end_date = user_date_input
                    df = df.loc[df[column].between(start_date, end_date)]
            elif is_categorical_dtype(df[column]) or df[column].unique().shape[0]<100:
                arr=df[column].unique()
                user_cat_input = right.multiselect(
                    f"Valueur de {column}",
                    arr
                    ,
                    default=list(arr),
                )
                df = df[df[column].isin(user_cat_input)]
            else:
                user_text_input = right.text_input(
                    f"Substring or regex in {column}",
                )
                if user_text_input:
                    df = df[df[column].astype(str).str.contains(user_text_input)]

    return df

df_perso=filter_dataframe(transf_df(df))
st.dataframe(df_perso)

st.sidebar.subheader("PARAMETRES DE LA BASE DE DONNEES")
df_selected=st.sidebar.radio("***:grey[Choisissez la base de données sur laquelle vous souhaitez réaliser les graphiques]***",['Base de données locale', 'Base de données personnalisée'])

if df_selected=='Base de données personnalisée':
    df=df_perso
else:
    df=df

# SECTION GRAPHIQUE

#Analyse univariée
st.header("ANALYSE GRAPHIQUE SUR UNE SEULE VARIABLE", divider='rainbow')
# Histogramme et Camembert sur la même ligne
cam, hist = st.columns(2,gap='medium')

with cam:
    st.subheader("Camembert")
    selected_categorical_variable_p = st.selectbox("***Sélectionnez une variable catégorielle pour le camembert***", ['SEXE', 'SITUATION MATRIMONIALE', "SECTEUR D'ACTIVITE", 'PROFESSION', 'AGENCES',"CHARGE D'AFFAIRES","TYPE DE PRÊT",'STATUT DU PRÊT'], index=1)
    category_counts = df[selected_categorical_variable_p].value_counts()
    fig_pie = px.pie(names=category_counts.index, values=category_counts.values, title=f"Répartition de la variable {selected_categorical_variable_p}",color_discrete_sequence=colors)
    fig_pie.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.25)
    st.plotly_chart(fig_pie, use_container_width=True)

with hist:
    st.subheader("Histogramme")
    selected_categorical_variable = st.selectbox("***Sélectionnez la variable catégorielle pour l'histogramme***", ['SEXE', 'SITUATION MATRIMONIALE', "SECTEUR D'ACTIVITE", 'PROFESSION', 'AGENCES',"CHARGE D'AFFAIRES","TYPE DE PRÊT","MOIS D'OCTROI DU CREDIT","ANNEE D'OCTROI DU CREDIT",'STATUT DU PRÊT'], index=6)
    fig_histogram = px.histogram(df, x=df[selected_categorical_variable], color=df[selected_categorical_variable],title=f"Histogramme de {selected_categorical_variable}",color_discrete_sequence=colors)
    fig_histogram.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.35)
    fig_histogram.update_traces( textfont_color='rgba(255, 255, 255, 1)')
    if selected_categorical_variable=="Mois":
        fig_histogram.update_xaxes(categoryorder='array', categoryarray=order_of_months)
    fig_histogram.update_xaxes(showticklabels=False)
    st.plotly_chart(fig_histogram,use_container_width=True)

# Section des analyses croisées
st.header("ANALYSES GRAPHIQUES CROISEES ENTRE DEUX VARIABLES", divider='rainbow')
quant,qual=st.columns(2,gap='medium')

    
with quant:
    st.subheader("Analyse croisée entre variables catégorielles")
    int_columns = df.select_dtypes(include="int").columns
    float_columns = df.select_dtypes(include="float").columns
    selected_variable_3 = st.selectbox("***Variable 1***", int_columns.union(float_columns))
    selected_variable_4 = st.selectbox("***Variable 2***",int_columns.union(float_columns),index=2)
    occurrences=df.groupby([selected_variable_3, selected_variable_4]).size().reset_index(name='count')
    #occurrences['Nombre de prêts risqués'] = (occurrences[selected_variable_3]+ occurrences[selected_variable_4])/2
    fig_scatter_matrix = px.scatter(occurrences, x=selected_variable_3, y=selected_variable_4,color_discrete_sequence=colors)
    fig_scatter_matrix.update_layout(title=f'Nuage de points entre {selected_variable_3} et {selected_variable_4}')
    fig_scatter_matrix.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.15)
    st.plotly_chart(fig_scatter_matrix, use_container_width=True)
with qual:
    #Type de l'histogramme croisé
    def barmode_selected(t):
        if t =='empilé':
            a='relative'  
        else: 
            a='group'
        return a
    
    st.subheader("Analyse croisée entre variables catégorielles")
    selected_variable_1 = st.selectbox("***Variable 1***", ['SEXE', 'SITUATION MATRIMONIALE', "SECTEUR D'ACTIVITE", 'PROFESSION', 'AGENCES',"CHARGE D'AFFAIRES","TYPE DE PRÊT","MOIS D'OCTROI DU CREDIT","ANNEE D'OCTROI DU CREDIT",'STATUT DU PRÊT'], index=4)
    selected_variable_2 = st.selectbox("***Variable 2***", ['SEXE', 'SITUATION MATRIMONIALE', "SECTEUR D'ACTIVITE", 'PROFESSION', 'AGENCES',"CHARGE D'AFFAIRES","TYPE DE PRÊT","MOIS D'OCTROI DU CREDIT","ANNEE D'OCTROI DU CREDIT",'STATUT DU PRÊT','AGE','DUREE DE REMBOURSEMENT',"TAUX D'INTERET","MONTANT SOLLICITE"],index=1)
    st.sidebar.write(" ")
    st.sidebar.write(" ")
    st.sidebar.subheader("PARAMETRES DES GRAPHIQUES")
    type_graph=st.sidebar.radio("***:grey[Choisissez le type d'histogramme croisé]***", ['empilé','étalé'])
    if selected_variable_2 in ['AGE','DUREE DE REMBOURSEMENT',"TAUX D'INTERET","MONTANT SOLLICITE"]:
        fig_croisé = px.bar(df.groupby(selected_variable_1)[selected_variable_2].mean().reset_index(), x=selected_variable_1,y=selected_variable_2, color=selected_variable_2,barmode=barmode_selected(type_graph))
    else:
        fig_croisé = px.bar(df, x=selected_variable_1, color=selected_variable_2,barmode=barmode_selected(type_graph),color_discrete_sequence= colors)
        
        if selected_variable_1=="Mois" or selected_variable_2=="Mois":
            fig_croisé.update_xaxes(categoryorder='array', categoryarray=order_of_months)
    fig_croisé.update_layout(title=f'Graphique en barres groupées - {selected_variable_1 } vs {selected_variable_2 }')
    fig_croisé.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},title_x=0.20)

    st.plotly_chart(fig_croisé,use_container_width=True)

st.header("GRAPHIQUE CHRONOLOGIQUE",divider="rainbow" )
occurences_an=df["ANNEE D'OCTROI DU CREDIT"].value_counts()
df["Nombre de prêts sur l'année"]=occurences_an[df["ANNEE D'OCTROI DU CREDIT"]].values
occurences_mo=df["MOIS D'OCTROI DU CREDIT"].value_counts()
df["Nombre de prêts sur le mois"]=occurences_mo[df["MOIS D'OCTROI DU CREDIT"]].values
fig_ann = px.area(df, x="ANNEE D'OCTROI DU CREDIT", y="Nombre de prêts sur l'année", color="MOIS D'OCTROI DU CREDIT",line_group="Nombre de prêts sur le mois",color_discrete_sequence= colors,custom_data=[df["MOIS D'OCTROI DU CREDIT"],df["ANNEE D'OCTROI DU CREDIT"],df["Nombre de prêts sur l'année"],df['Nombre de prêts sur le mois']])
fig_ann.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.3)',},height= 500,width= 1280)
fig_ann.update_xaxes(categoryorder='array', categoryarray=order_of_years)
fig_ann.update_traces(hovertemplate='<b>Année</b>: %{customdata[1]}<br>'
                                    "<b>Nbre de prêts sur l'année</b>: %{customdata[2]}<br>"
                                    "<b>Mois</b>: %{customdata[0]}<br>"
                                    '<b>Nbre de prêts sur le mois</b>: %{customdata[3]}<br>',
                                    hoverlabel=dict(font=dict(size=16, color='white'))),
#st.plotly_chart(fig_ann)
st.header("REPRESENTATION CARTOGRAPHIQUE DU RISQUE DE DEFAUT DE PAIEMENT",divider="rainbow" )
#Map_agencies
data = openpyxl.load_workbook(r"C:\Users\DELL I5\Desktop\Classeur2.xlsx")
datas = data.active
donnees = []
for ligne in datas.iter_rows(values_only=True):
    donnees.append(list(ligne))
en_tetes = donnees[0]
donnees = donnees[1:]
df2 = pd.DataFrame(donnees, columns=en_tetes)
df= pd.merge(df, df2, on='AGENCES')

occurences=df['AGENCES'].value_counts()
df['size']=occurences[df['AGENCES']].values

d=df.groupby('AGENCES')['STATUT'].sum()
df['Nbre de crédits risqués par agence']=d[df['AGENCES']].values
df["Taux de défauts de paiement"]=df['Nbre de crédits risqués par agence']/df['size']
df["Taux de défauts de paiement en pct"]=df["Taux de défauts de paiement"].map(lambda x: f'{x:.2%}')
center_lat = 5.31908
center_lon = -4.01299
px.set_mapbox_access_token('pk.eyJ1IjoiZXhhbXBsZXMiLCJhIjoiY2lqbmpqazdlMDBsdnRva284cWd3bm11byJ9.V6Hg2oYJwMAxeoR9GEzkAA')

fig_map = px.scatter_mapbox(df, lat=df.latitude, lon=df.longitude, color="Taux de défauts de paiement", size="size",
                  color_continuous_scale=px.colors.diverging.RdYlGn_r, zoom=12,custom_data=[df['AGENCES'],df['size'],df['Taux de défauts de paiement en pct']],size_max=40)

fig_map.update_layout({'plot_bgcolor': 'rgba(0, 0, 0, 0)','paper_bgcolor': 'rgba(0, 0, 0, 0.05)','height': 700,'width': 1540,'mapbox': {'center': {'lat': center_lat, 'lon': center_lon}, 'style':'dark'}})
fig_map.update_traces(text=df['Taux de défauts de paiement en pct'], hovertemplate='<b>Agence</b>: %{customdata[0]}<br>'
                                                                                      '<b>Latitude</b>: %{lat:.4f}<br>'
                                                                                      '<b>Longitude</b>: %{lon:.4f}<br>'
                                                                                      '<b>Nbre de prêts</b>: %{customdata[1]}<br>'
                                                                                      '<b>Taux de prêts en défaut de paiement</b>: %{customdata[2]}',
                                                                                      hoverlabel=dict(font=dict(size=16, color='white')),

                                                                                      )

st.plotly_chart(fig_map)
